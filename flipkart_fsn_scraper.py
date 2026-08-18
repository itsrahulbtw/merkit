"""Add Flipkart product links to an Excel file and download product gallery images.

Usage:
    python flipkart_fsn_scraper.py fsn.xlsx

The input file is never changed. Results are written to output/YYYY-MM-DD/.
"""

from __future__ import annotations

import argparse
import mimetypes
import re
import time
from datetime import date
from pathlib import Path
from urllib.parse import urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font


PRODUCT_URL = "https://www.flipkart.com/product/p/itme?pid={fsn}"
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0 Safari/537.36"
    ),
    "Accept-Language": "en-IN,en;q=0.9",
}


def normalise_fsn(value: object) -> str:
    """Return a clean FSN, or an empty string for a blank cell."""
    return str(value or "").strip().upper()


def image_extension(response: requests.Response, url: str) -> str:
    content_type = response.headers.get("Content-Type", "").split(";", 1)[0]
    if content_type == "image/jpeg":
        return ".jpg"
    if content_type == "image/png":
        return ".png"
    if content_type == "image/webp":
        return ".webp"
    return Path(urlparse(url).path).suffix or mimetypes.guess_extension(content_type) or ".jpg"


def hero_image_urls(page_html: str) -> list[str]:
    """Extract gallery images, excluding small thumbnails and recommendation cards.

    Flipkart's gallery uses high-resolution ``rukminim*.flixcart.com/image`` URLs;
    product suggestions use smaller 80/530 pixel variants. We retain only the
    high-resolution gallery URLs and de-duplicate them while preserving order.
    """
    soup = BeautifulSoup(page_html, "html.parser")
    urls: list[str] = []
    for image in soup.select('img[src*="rukminim"][src*="/image/"]'):
        url = image.get("src", "")
        # Main product gallery images are served at 800px or wider.
        match = re.search(r"/image/(\d+)/", url)
        if not match or int(match.group(1)) < 800:
            continue
        if url not in urls:
            urls.append(url)
    return urls


def download_images(session: requests.Session, fsn: str, url: str, image_dir: Path) -> tuple[int, str]:
    response = session.get(url, timeout=40)
    response.raise_for_status()
    urls = hero_image_urls(response.text)
    if not urls:
        return 0, "No hero images found"

    fsn_dir = image_dir / fsn
    fsn_dir.mkdir(parents=True, exist_ok=True)
    saved = 0
    for index, image_url in enumerate(urls, start=1):
        image_response = session.get(image_url, timeout=40)
        image_response.raise_for_status()
        if not image_response.headers.get("Content-Type", "").startswith("image/"):
            raise ValueError(f"Unexpected download type: {image_response.headers.get('Content-Type')}")
        destination = fsn_dir / f"hero_{index:02d}{image_extension(image_response, image_url)}"
        destination.write_bytes(image_response.content)
        saved += 1
    return saved, "Scraped"


def ensure_column(ws, header: str) -> int:
    for cell in ws[1]:
        if str(cell.value or "").strip().lower() == header.lower():
            return cell.column
    column = ws.max_column + 1
    ws.cell(1, column, header).font = Font(bold=True)
    return column


def process_workbook(input_file: Path, output_root: Path, delay: float) -> Path:
    today_folder = output_root / date.today().isoformat()
    image_dir = today_folder / "images"
    today_folder.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(input_file)
    session = requests.Session()
    session.headers.update(HEADERS)

    for ws in workbook.worksheets:
        headers = {str(cell.value or "").strip().lower(): cell.column for cell in ws[1]}
        fsn_column = headers.get("fsn")
        if not fsn_column:
            print(f"Skipping '{ws.title}': no 'fsn' header found.")
            continue

        link_column = ensure_column(ws, "link")
        status_column = ensure_column(ws, "scrape status")
        count_column = ensure_column(ws, "hero images downloaded")
        folder_column = ensure_column(ws, "image folder")
        detail_column = ensure_column(ws, "scrape detail")

        for row in range(2, ws.max_row + 1):
            fsn = normalise_fsn(ws.cell(row, fsn_column).value)
            if not fsn:
                continue
            product_url = PRODUCT_URL.format(fsn=fsn)
            link_cell = ws.cell(row, link_column, product_url)
            link_cell.hyperlink = product_url
            link_cell.style = "Hyperlink"
            try:
                saved, detail = download_images(session, fsn, product_url, image_dir)
                status = "Scraped" if saved else "Not scraped"
                ws.cell(row, status_column, status)
                ws.cell(row, count_column, saved)
                ws.cell(row, folder_column, str((image_dir / fsn).resolve()) if saved else "")
                ws.cell(row, detail_column, detail)
                print(f"{fsn}: {status} ({saved} image(s))")
            except (requests.RequestException, ValueError) as error:
                ws.cell(row, status_column, "Not scraped")
                ws.cell(row, count_column, 0)
                ws.cell(row, folder_column, "")
                ws.cell(row, detail_column, str(error)[:300])
                print(f"{fsn}: Not scraped ({error})")
            time.sleep(delay)

    output_file = today_folder / f"{input_file.stem}_scraped.xlsx"
    workbook.save(output_file)
    return output_file


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("excel_file", type=Path, help="Excel .xlsx file with an 'fsn' column")
    parser.add_argument("--output", type=Path, default=Path("output"), help="Output root (default: output)")
    parser.add_argument("--delay", type=float, default=1.0, help="Seconds between product pages (default: 1)")
    args = parser.parse_args()
    if not args.excel_file.is_file():
        parser.error(f"File not found: {args.excel_file}")
    print(f"Saved workbook: {process_workbook(args.excel_file, args.output, args.delay).resolve()}")


if __name__ == "__main__":
    main()
